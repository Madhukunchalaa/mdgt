from django.shortcuts import render
import csv
import json
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.apps import apps
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from Employee.models import Employee


# -------------------------------------------------------------------
# Get model dynamically
# -------------------------------------------------------------------
def get_model_by_name(model_name):
    # Normalize model name
    normalized_name = model_name.lower().strip()
    
    # Handle common name variations
    name_mapping = {
        "matgattribute": "matgattributeitem",
        "materialattribute": "matgattributeitem",
        "matgattributeitem": "matgattributeitem",
        "itemmasterold": "itemmaster",
    }
    
    # Check if we have a mapping
    if normalized_name in name_mapping:
        normalized_name = name_mapping[normalized_name]
    
    # Search for exact match first
    for model in apps.get_models():
        if model.__name__.lower() == normalized_name:
            return model
    
    # If not found, try partial match (e.g., "MatgAttribute" matches "MatgAttributeItem")
    for model in apps.get_models():
        if normalized_name in model.__name__.lower() or model.__name__.lower() in normalized_name:
            return model
    
    return None


# -------------------------------------------------------------------
# Convert CSV/JSON/Excel value → correct Python type
# -------------------------------------------------------------------
def convert_value(field, value):
    if value is None or value == "":
        return None

    internal_type = field.get_internal_type()

    try:
        if internal_type in ["IntegerField", "BigIntegerField"]:
            return int(value)

        if internal_type == "BooleanField":
            return str(value).lower() in ["1", "true", "yes"]

        if internal_type == "FloatField":
            return float(value)

        if internal_type == "DateField":
            return timezone.datetime.fromisoformat(value).date()

        if internal_type == "DateTimeField":
            return timezone.datetime.fromisoformat(value)

        if internal_type == "ForeignKey":
            # FKs must match to_field, not always "id"
            return field.related_model.objects.get(**{field.target_field.name: value})

        return value

    except Exception:
        return value  # safe fallback


# -------------------------------------------------------------------
# Generate Upload Log (Excel with Status & Error)
# -------------------------------------------------------------------
def generate_upload_log(original_data, results, model_name):
    """
    original_data: List of dicts (the original upload)
    results: List of dicts matching original_data length with {'status': 'Success'|'Error', 'error': 'msg'}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload Results"

    if not original_data:
        ws.cell(row=1, column=1, value="No data uploaded")
        return wb

    # Get all unique headers from original data
    headers = []
    for row in original_data:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
    
    # Add status columns at the beginning
    all_headers = ["Upload Status", "Error Details"] + headers

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    # Write data rows
    for row_idx, (orig_row, res) in enumerate(zip(original_data, results), start=2):
        # Status column
        status = res.get('status', 'Unknown')
        status_cell = ws.cell(row=row_idx, column=1, value=status)
        if status == "Success":
            status_cell.font = Font(color="008000") # Green
        elif status == "Error":
            status_cell.font = Font(color="FF0000") # Red

        # Error details column
        ws.cell(row=row_idx, column=2, value=res.get('error', ''))

        # Original data columns
        for col_idx, header in enumerate(headers, start=3):
            ws.cell(row=row_idx, column=col_idx, value=orig_row.get(header, ""))

    return wb


# -------------------------------------------------------------------
# Handler: ItemMaster Phase 1 — Insert Base Data
# -------------------------------------------------------------------
def handle_itemmaster_phase_1(data, request):
    from itemmaster.models import ItemMaster
    from MaterialType.models import MaterialType
    from matgroups.models import MatGroup

    now = timezone.now()
    objs = []
    errors = []

    for idx, row in enumerate(data):
        try:
            # Helper function to get value by multiple possible keys (handles different header formats)
            def get_value(row, possible_keys):
                for key in possible_keys:
                    if key in row and row[key] and str(row[key]).strip():
                        return str(row[key]).strip()
                return None
            
            # Get mat_type_code (handle different header formats)
            mat_type_code_value = get_value(row, ["mat_type_code", "Mat Type Code", "mat type code", "MAT_TYPE_CODE"])
            mat_type_code = None
            if mat_type_code_value:
                mat_type_code = MaterialType.objects.filter(mat_type_code=mat_type_code_value).first()
                if not mat_type_code:
                    errors.append({"row": idx + 2, "error": f"MaterialType '{mat_type_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx + 2, "error": "mat_type_code is required"})
                continue
            
            # Get mgrp_code (handle different header formats)
            mgrp_code_value = get_value(row, ["mgrp_code", "Mgrp Code", "mgrp code", "MGRP_CODE"])
            mgrp_code = None
            if mgrp_code_value:
                mgrp_code = MatGroup.objects.filter(mgrp_code=mgrp_code_value).first()
                if not mgrp_code:
                    errors.append({"row": idx + 2, "error": f"MatGroup '{mgrp_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx + 2, "error": "mgrp_code is required"})
                continue
            
            # Get short_name (optional, max 40 chars)
            short_name = get_value(row, ["short_name", "Short Name", "short name", "SHORT_NAME"]) or ""
            if short_name and len(short_name) > 40:
                errors.append({"row": idx + 2, "error": f"short_name exceeds 40 characters (got {len(short_name)})"})
                continue
            
            # Convert sap_item_id to int if it's a string
            sap_item_id_value = get_value(row, ["sap_item_id", "Sap Item Id", "SAP Item Id", "SAP Item ID", "sap item id", "SAP_ITEM_ID"])
            sap_item_id = None
            if sap_item_id_value:
                try:
                    sap_item_id = int(float(sap_item_id_value))
                except (ValueError, TypeError):
                    sap_item_id = None

            # Get other optional fields
            long_name    = get_value(row, ["long_name", "Long Name", "long name", "LONG_NAME"])
            mgrp_long_name = get_value(row, ["mgrp_long_name", "Mgrp Long Name", "mgrp long name", "MGRP_LONG_NAME"])
            sap_name     = get_value(row, ["sap_name", "Sap Name", "SAP Name", "sap name", "SAP_NAME"])
            search_text  = get_value(row, ["search_text", "Search Text", "search text", "SEARCH_TEXT"])

            # New specification fields
            item_type   = get_value(row, ["item_type", "Type", "TYPE"])
            item_number = get_value(row, ["item_number", "Number", "NUMBER"])
            moc         = get_value(row, ["moc", "Moc", "MOC"])
            item_size   = get_value(row, ["item_size", "Size", "SIZE"])
            part_number = get_value(row, ["part_number", "Part Number", "part number", "PART_NUMBER", "PART NUMBER"])
            model       = get_value(row, ["model", "Model", "MODEL"])
            make        = get_value(row, ["make", "Make", "MAKE"])

            # Build attributes JSON by matching spec fields to MatgAttributeItem names (case-insensitive)
            spec_map = {
                "type": item_type, "number": item_number, "moc": moc,
                "size": item_size, "part number": part_number, "model": model, "make": make,
            }
            attributes = {}
            from matg_attributes.models import MatgAttributeItem
            for attr in MatgAttributeItem.objects.filter(mgrp_code=mgrp_code, is_deleted=False):
                val = spec_map.get(attr.attribute_name.lower().strip())
                if val:
                    attributes[attr.attribute_name] = val

            # Auto-generate short_name from sap_name + spec fields if not provided
            if not short_name:
                parts = [p for p in [sap_name, item_type, item_number, moc, item_size, part_number, model, make] if p]
                short_name = ", ".join(parts)[:40]

            fields = dict(
                mat_type_code=mat_type_code,
                mgrp_code=mgrp_code,
                short_name=short_name,
                long_name=long_name,
                mgrp_long_name=mgrp_long_name,
                sap_name=sap_name,
                search_text=search_text,
                attributes=attributes,
                item_type=item_type,
                item_number=item_number,
                moc=moc,
                item_size=item_size,
                part_number=part_number,
                model=model,
                make=make,
                updated=now,
            )

            # Update existing record if sap_item_id matches, else create new
            if sap_item_id:
                existing = ItemMaster.objects.filter(sap_item_id=sap_item_id).first()
                if existing:
                    for k, v in fields.items():
                        setattr(existing, k, v)
                    existing.save()
                    objs.append(existing)
                    continue

            objs.append(ItemMaster(sap_item_id=sap_item_id, created=now, **fields))

        except Exception as e:
            import traceback
            errors.append({"row": idx + 2, "error": f"{str(e)}"})

    new_objs = [o for o in objs if not o.pk]
    if new_objs:
        ItemMaster.objects.bulk_create(new_objs, ignore_conflicts=True)

    # Populate final row_results (simplified for bulk_create)
    # Since bulk_create with ignore_conflicts=True is used, we assume all valid-looking rows were "Success" 
    # unless they hit a hard error caught in the loop.
    row_results = []
    errors_dict = {e['row']: e['error'] for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": "ItemMaster Phase 1 upload complete",
        "inserted": len(objs),
        "errors": errors,
        "row_results": row_results
    }


# -------------------------------------------------------------------
# Validation Functions
# -------------------------------------------------------------------
def validate_attribute_value(value, validation_type):
    """
    Validate attribute value based on validation type.
    Returns (is_valid, error_message)
    """
    if not validation_type or not value:
        return True, None
    
    validation_type = validation_type.lower().strip()
    
    if validation_type == "alpha":
        if not value.replace(" ", "").isalpha():
            return False, f"Value '{value}' must contain only alphabetic characters"
    
    elif validation_type == "numeric":
        if not value.replace(".", "").replace("-", "").isdigit():
            return False, f"Value '{value}' must be numeric"
    
    elif validation_type == "alphanumeric":
        if not value.replace(" ", "").isalnum():
            return False, f"Value '{value}' must contain only alphanumeric characters"
    
    elif validation_type == "wholenumber":
        try:
            num = float(value)
            if num < 0 or num != int(num):
                return False, f"Value '{value}' must be a whole number (non-negative integer)"
        except ValueError:
            return False, f"Value '{value}' must be a whole number"
    
    elif validation_type == "integer":
        try:
            int(value)
        except ValueError:
            return False, f"Value '{value}' must be an integer"
    
    elif validation_type == "decimal":
        try:
            float(value)
        except ValueError:
            return False, f"Value '{value}' must be a decimal number"
    
    return True, None


# -------------------------------------------------------------------
# Handler: ItemMaster Phase 2 — Merge Attributes JSON
# -------------------------------------------------------------------
def handle_itemmaster_phase_2(data, request):
    from itemmaster.models import ItemMaster
    from itemmaster.views import format_short_name
    from matg_attributes.models import MatgAttributeItem
    import json

    updated = 0
    created = 0
    unchanged = 0
    errors = []

    # Helper function to get value by multiple possible keys (handles different header formats)
    def get_value(row, possible_keys):
        for key in possible_keys:
            if key in row and row[key] is not None and str(row[key]).strip():
                return str(row[key]).strip()
        return None

    # Detect format: wide (no "Attribute Name" column) vs vertical (legacy)
    sample_row = data[0] if data else {}
    fixed_cols = {"sap item id", "sap_item_id", "uom", "sap item id", "sap material number"}
    is_wide_format = not any(
        k.lower().strip() in ("attribute name", "attribute_name", "attribute value", "attribute_value")
        for k in sample_row.keys()
    )

    if is_wide_format:
        # Wide format: each column (except Sap Item Id & Uom) is an attribute name
        for idx, row in enumerate(data, start=2):
            try:
                sap_item_id_value = get_value(row, ["sap_item_id", "Sap Item Id", "SAP Item Id", "SAP Item ID", "sap item id", "SAP_ITEM_ID"])
                if not sap_item_id_value:
                    errors.append({"row": idx, "error": "sap_item_id is required"})
                    continue
                try:
                    sap = int(float(sap_item_id_value))
                except (ValueError, TypeError):
                    errors.append({"row": idx, "error": f"Invalid sap_item_id: {sap_item_id_value}"})
                    continue

                item = ItemMaster.objects.filter(sap_item_id=sap).first()
                if not item:
                    errors.append({"row": idx, "error": f"ItemMaster with sap_item_id {sap} not found"})
                    continue

                uom = get_value(row, ["uom", "Uom", "UOM"])
                attributes = item.attributes or {}
                if isinstance(attributes, str):
                    try:
                        attributes = json.loads(attributes)
                    except json.JSONDecodeError:
                        attributes = {}

                skip_cols = {"sap item id", "sap_item_id", "uom", "sap material number"}
                for col_key, col_val in row.items():
                    if col_key.lower().strip() in skip_cols:
                        continue
                    attr_name = col_key.strip()
                    attr_value = str(col_val).strip() if col_val is not None and str(col_val).strip() else ""

                    def norm(s): return s.lower().replace(" ", "")
                    existing_key = next((k for k in attributes if norm(k) == norm(attr_name)), None)
                    store_key = existing_key if existing_key else attr_name

                    old_attr_data = attributes.get(store_key)
                    old_value = None
                    if isinstance(old_attr_data, dict):
                        old_value = old_attr_data.get("value", "")
                    elif old_attr_data is not None:
                        old_value = str(old_attr_data)

                    if uom:
                        attributes[store_key] = {"value": attr_value, "uom": uom}
                    else:
                        attributes[store_key] = attr_value

                    if old_value is None or old_value == "":
                        created += 1
                    elif old_value != attr_value:
                        updated += 1
                    else:
                        unchanged += 1

                item.attributes = attributes
                rebuilt = format_short_name(item.sap_name, attributes)
                if rebuilt and len(rebuilt) >= 3:
                    item.short_name = rebuilt[:40]
                item.save()

            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        # Row results for wide format
        row_results = []
        errors_dict = {e['row']: e['error'] for e in errors}
        for i in range(len(data)):
            row_num = i + 2
            if row_num in errors_dict:
                row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
            else:
                row_results.append({'status': 'Success', 'error': ''})

        return {
            "message": "ItemMaster Phase 2 attribute merge complete",
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "errors": errors,
            "row_results": row_results
        }

    # --- Legacy vertical format below ---
    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            # Get sap_item_id (handle different header formats)
            sap_item_id_value = get_value(row, ["sap_item_id", "Sap Item Id", "SAP Item Id", "SAP Item ID", "sap item id", "SAP_ITEM_ID"])
            if not sap_item_id_value:
                errors.append({"row": idx, "error": "sap_item_id is required"})
                continue

            # Convert to int
            try:
                sap = int(float(sap_item_id_value))
            except (ValueError, TypeError):
                errors.append({"row": idx, "error": f"Invalid sap_item_id: {sap_item_id_value}"})
                continue

            # Find item
            item = ItemMaster.objects.filter(sap_item_id=sap).first()
            if not item:
                errors.append({"row": idx, "error": f"ItemMaster with sap_item_id {sap} not found"})
                continue

            # Get attribute_name (handle different header formats)
            attr_name = get_value(row, ["attribute_name", "Attribute Name", "attribute name", "ATTRIBUTE_NAME"])
            if not attr_name:
                errors.append({"row": idx, "error": "attribute_name is required"})
                continue

            # Get attribute_value (handle different header formats)
            attr_value = get_value(row, ["attribute_value", "Attribute Value", "attribute value", "ATTRIBUTE_VALUE"])
            if attr_value is None:  # Allow empty string but not None
                attr_value = ""

            # Get UOM (handle different header formats) - optional
            uom = get_value(row, ["uom", "Uom", "UOM", "Unit Of Measure", "unit of measure"])

            # Look up the attribute definition to get validation rules
            attr_def = None
            try:
                attr_def = MatgAttributeItem.objects.filter(
                    mgrp_code=item.mgrp_code,
                    attribute_name=attr_name,
                    is_deleted=False
                ).first()
            except Exception as e:
                # If we can't find the attribute definition, we'll skip validation
                pass

            # Only validate if a regex/type validation rule is explicitly set; skip possible_values check on bulk upload
            if attr_def and attr_def.validation and attr_value:
                is_valid, error_msg = validate_attribute_value(attr_value, attr_def.validation)
                if not is_valid:
                    errors.append({"row": idx, "error": error_msg})
                    continue

            # Ensure JSON is dict
            attributes = item.attributes or {}
            if isinstance(attributes, str):
                try:
                    attributes = json.loads(attributes)
                except json.JSONDecodeError:
                    attributes = {}

            # Get old value before updating (for tracking changes)
            # Case+space insensitive key match to avoid duplicates on re-upload
            def norm(s): return s.lower().replace(" ", "")
            existing_key = next((k for k in attributes if norm(k) == norm(attr_name)), None)
            store_key = existing_key if existing_key else attr_name

            old_attr_data = attributes.get(store_key)
            old_value = None
            if isinstance(old_attr_data, dict):
                old_value = old_attr_data.get("value", "")
            elif old_attr_data is not None:
                old_value = str(old_attr_data)

            # Store attribute value (with UOM if provided)
            if uom:
                attributes[store_key] = {"value": attr_value, "uom": uom}
            else:
                attributes[store_key] = attr_value

            # Track changes
            if old_value is None or old_value == "":
                created += 1
            elif old_value != attr_value:
                updated += 1
            else:
                unchanged += 1

            # Update item attributes and rebuild short_name
            item.attributes = attributes
            rebuilt = format_short_name(item.sap_name, attributes)
            if rebuilt and len(rebuilt) >= 3:
                item.short_name = rebuilt[:40]
            item.save()

        except Exception as e:
            import traceback
            errors.append({"row": idx, "error": f"{str(e)}"})

    # Row results for legacy vertical format
    row_results = []
    errors_dict = {e['row']: e['error'] for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": "ItemMaster Phase 2 attribute merge complete",
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "errors": errors,
        "row_results": row_results
    }


# -------------------------------------------------------------------
# Generic Handler: For any model that doesn't have a specific handler
# -------------------------------------------------------------------
def handle_generic_model_upload(data, request, Model, model_name):
    """
    Generic handler for bulk uploading any model.
    Automatically handles ForeignKey fields and converts data types.
    """
    now = timezone.now()
    objs = []
    errors = []
    
    # Get the user from request if available (for audit fields)
    user = None
    # Try to get user from token if available
    try:
        from Employee.decorator import get_user_from_token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if token:
            user = get_user_from_token(token)
    except:
        pass
    
    # Get all model fields (only concrete fields, not reverse relations)
    model_fields = {}
    for field in Model._meta.concrete_fields:
        model_fields[field.name] = field
    
    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            obj_data = {}
            
            # Process each field in the row
            for field_name, value in row.items():
                # Normalize field name (handle spaces, case differences)
                normalized_field_name = field_name.lower().replace(' ', '_').replace('-', '_').strip()
                
                # Find matching field (case-insensitive)
                matching_field = None
                for model_field_name, model_field in model_fields.items():
                    if model_field_name.lower() == normalized_field_name:
                        matching_field = model_field
                        break
                
                if not matching_field:
                    # Field not found in model, skip it (might be a header formatting issue)
                    continue
                
                # Skip audit fields (they'll be set automatically)
                if matching_field.name in ['id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted']:
                    continue
                
                # Handle ForeignKey fields
                if hasattr(matching_field, 'related_model') and matching_field.related_model:
                    if value and str(value).strip():
                        # Get the related model
                        related_model = matching_field.related_model
                        # Try to find by primary key or by the target field
                        target_field = matching_field.target_field.name if hasattr(matching_field, 'target_field') else 'pk'
                        
                        try:
                            # Try to get by the target field value
                            fk_obj = related_model.objects.filter(**{target_field: str(value).strip()}).first()
                            if fk_obj:
                                obj_data[matching_field.name] = fk_obj
                            else:
                                errors.append({
                                    "row": idx,
                                    "field": field_name,
                                    "error": f"Foreign key value '{value}' not found in {related_model.__name__}"
                                })
                        except Exception as e:
                            errors.append({
                                "row": idx,
                                "field": field_name,
                                "error": f"Error resolving foreign key: {str(e)}"
                            })
                else:
                    # Handle regular fields
                    try:
                        converted_value = convert_value(matching_field, value)
                        obj_data[matching_field.name] = converted_value
                    except Exception as e:
                        errors.append({
                            "row": idx,
                            "field": field_name,
                            "error": f"Error converting value: {str(e)}"
                        })
            
            # Set audit fields if they exist
            if 'created' in model_fields:
                obj_data['created'] = now
            if 'updated' in model_fields:
                obj_data['updated'] = now
            if 'createdby' in model_fields and user:
                obj_data['createdby'] = user
            if 'updatedby' in model_fields and user:
                obj_data['updatedby'] = user
            
            # Create the object
            if obj_data:
                obj = Model(**obj_data)
                objs.append(obj)
        
        except Exception as e:
            errors.append({
                "row": idx,
                "error": f"Error processing row: {str(e)}"
            })
    
    # Bulk create objects
    if objs:
        try:
            Model.objects.bulk_create(objs, ignore_conflicts=True)
        except Exception as e:
            return JsonResponse({
                "error": f"Bulk create failed: {str(e)}",
                "errors": errors
            }, status=400)
    
    # Populate final row_results
    row_results = []
    errors_dict = {e['row']: e.get('error', 'Error') for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": f"{model_name} upload complete",
        "inserted": len(objs),
        "errors": errors,
        "row_results": row_results
    }


# -------------------------------------------------------------------
# Handler: MatgAttributeItem Phase 1 — Insert Allowed Values + UOMs
# -------------------------------------------------------------------
def handle_matgattribute_phase_1(data, request):
    from matg_attributes.models import MatgAttributeItem
    from matgroups.models import MatGroup

    inserted = 0
    updated = 0
    errors = []

    # Helper function to get value by multiple possible keys (handles different header formats)
    def get_value(row, possible_keys):
        for key in possible_keys:
            if key in row and row[key] and str(row[key]).strip():
                return str(row[key]).strip()
        return None

    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            # Get mgrp_code (handle different header formats)
            mgrp_code_value = get_value(row, ["mgrp_code", "Mgrp Code", "mgrp code", "MGRP_CODE"])
            mgrp_code = None
            if mgrp_code_value:
                mgrp_code = MatGroup.objects.filter(mgrp_code=mgrp_code_value).first()
                if not mgrp_code:
                    errors.append({"row": idx, "error": f"MatGroup '{mgrp_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx, "error": "mgrp_code is required"})
                continue

            # Get attribute_name (handle different header formats)
            attribute_name = get_value(row, ["attribute_name", "Attribute Name", "attribute name", "ATTRIBUTE_NAME"])
            if not attribute_name:
                errors.append({"row": idx, "error": "attribute_name is required"})
                continue

            # Get possible_values (handle different header formats)
            possible_values_str = get_value(row, ["possible_values", "Possible Values", "possible values", "POSSIBLE_VALUES"])
            possible_vals = []
            if possible_values_str:
                possible_vals = [
                    x.strip() for x in possible_values_str.split(",")
                    if x.strip()
                ]

            # Get uom (handle different header formats)
            uom_str = get_value(row, ["uom", "Uom", "UOM", "Unit Of Measure"])
            uom = uom_str if uom_str else None

            # Get print_priority (handle different header formats)
            print_priority_str = get_value(row, ["print_priority", "Print Priority", "print priority", "PRINT_PRIORITY"])
            print_priority = None
            if print_priority_str:
                try:
                    print_priority = int(float(print_priority_str))
                except (ValueError, TypeError):
                    print_priority = None

            # Get validation (handle different header formats)
            validation = get_value(row, ["validation", "Validation", "VALIDATION"])

            # update_or_create so edited rows are reflected in DB
            _, created = MatgAttributeItem.objects.update_or_create(
                mgrp_code=mgrp_code,
                attribute_name=attribute_name,
                defaults={
                    "possible_values": possible_vals,
                    "uom": uom,
                    "print_priority": print_priority,
                    "validation": validation,
                    "is_deleted": False,
                }
            )
            if created:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            errors.append({"row": idx, "error": f"{str(e)}"})

    # Populate final row_results
    row_results = []
    errors_dict = {e['row']: e['error'] for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": "MatGroup Attribute Definitions imported",
        "inserted": inserted,
        "updated": updated,
        "errors": errors,
        "row_results": row_results
    }


# -------------------------------------------------------------------
# -------------------------------------------------------------------
# MatGroup Upload Handler
# -------------------------------------------------------------------
def handle_matgroup_upload(data, request):
    from matgroups.models import MatGroup
    from supergroups.models import SuperGroup

    VALID_SEARCH_TYPES = {"service", "Materials", "spares"}

    now = timezone.now()
    objs = []
    errors = []

    def get_val(row, keys):
        for k in keys:
            if k in row and str(row[k]).strip():
                return str(row[k]).strip()
        return None

    for idx, row in enumerate(data, start=2):
        try:
            mgrp_code = get_val(row, ["Mgrp Code", "mgrp_code", "MGRP_CODE"])
            sgrp_code_val = get_val(row, ["Sgrp Code", "sgrp_code", "SGRP_CODE"])
            raw_search_type = get_val(row, ["Search Type", "search_type"])
            shortname = get_val(row, ["Mgrp Shortname", "mgrp_shortname"])
            longname = get_val(row, ["Mgrp Longname", "mgrp_longname"])

            # Validate search_type — case-insensitive match against allowed DB choices
            search_type_map = {
                "service": "service",
                "materials": "Materials",
                "spares": "spares"
            }
            
            search_type = "Materials" # default
            if raw_search_type:
                clean_type = raw_search_type.strip().lower()
                if clean_type in search_type_map:
                    search_type = search_type_map[clean_type]

            # Cap field lengths to match model constraints
            if shortname:
                shortname = shortname[:150]
            if longname:
                longname = longname[:150]

            if not mgrp_code:
                errors.append({"row": idx, "error": "Mgrp Code is required"})
                continue

            existing = MatGroup.objects.filter(mgrp_code=mgrp_code).first()
            if existing:
                if not existing.is_deleted:
                    # Truly active duplicate — skip with error
                    errors.append({"row": idx, "error": f"MatGroup '{mgrp_code}' already exists"})
                    continue
                # Soft-deleted record — restore it with the uploaded data
                sgrp_obj = None
                if sgrp_code_val:
                    sgrp_key = sgrp_code_val[:5]
                    sgrp_obj, _ = SuperGroup.objects.get_or_create(
                        sgrp_code=sgrp_key,
                        defaults={"sgrp_name": sgrp_code_val, "dept_name": sgrp_code_val[:20]},
                    )
                existing.sgrp_code = sgrp_obj
                existing.search_type = search_type
                existing.mgrp_shortname = shortname
                existing.mgrp_longname = longname
                existing.is_deleted = False
                existing.updated = now
                existing.save()
                continue

            sgrp_obj = None
            if sgrp_code_val:
                # Truncate to max_length=5 to avoid DB validation errors
                sgrp_key = sgrp_code_val[:5]
                # Auto-create SuperGroup if it doesn't exist yet
                sgrp_obj, _ = SuperGroup.objects.get_or_create(
                    sgrp_code=sgrp_key,
                    defaults={"sgrp_name": sgrp_code_val, "dept_name": sgrp_code_val[:20]},
                )

            objs.append(MatGroup(
                mgrp_code=mgrp_code,
                sgrp_code=sgrp_obj,
                search_type=search_type,
                mgrp_shortname=shortname,
                mgrp_longname=longname,
                notes="",
                uom_values=[],
                created=now,
                updated=now,
            ))

        except Exception as e:
            errors.append({"row": idx, "error": f"Unexpected error: {str(e)}"})

    if objs:
        MatGroup.objects.bulk_create(objs, ignore_conflicts=True)

    # Populate final row_results
    row_results = []
    errors_dict = {e['row']: e['error'] for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": "MatGroup upload complete",
        "inserted": len(objs),
        "errors": errors,
        "row_results": row_results
    }


# MAIN BULK UPLOAD FUNCTION WITH PHASE ROUTING
# -------------------------------------------------------------------
@csrf_exempt
def bulk_upload(request):
    model_name = request.POST.get("model")
    phase = request.POST.get("phase", "1")

    if not model_name:
        return JsonResponse({"error": "Model name is required"}, status=400)

    Model = get_model_by_name(model_name)
    if not Model:
        return JsonResponse({"error": f"Invalid model: {model_name}"}, status=400)

    # -------------------------------------------------------------------
    # Parse file (CSV or Excel)
    # -------------------------------------------------------------------
    data = []
    file = request.FILES.get("file")

    if not file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    ext = file.name.split('.')[-1].lower()

    try:
        # Excel Upload
        if ext in ["xlsx", "xls"]:
            wb = openpyxl.load_workbook(file)
            
            # For ItemMaster phase 2, look for "Attributes" sheet
            if model_name and model_name.lower() == "itemmaster" and phase == "2":
                # Try to find Attributes sheet (case-insensitive)
                attributes_sheet = None
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower().strip() in ("attributes", "attribute settings"):
                        attributes_sheet = wb[sheet_name]
                        break
                
                if attributes_sheet:
                    sheet = attributes_sheet
                else:
                    # Fall back to active sheet if Attributes not found
                    sheet = wb.active
            else:
                # For phase 1 or other models, use active sheet
                sheet = wb.active
            
            header = [str(c.value).strip() if c.value else "" for c in next(sheet.rows)]  # Filter out None headers and convert to string
            
            # Remove empty headers
            header = [h for h in header if h]

            for row in sheet.iter_rows(min_row=2):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(header) and header[idx]:
                        cell_value = "" if cell.value is None else str(cell.value).strip()
                        row_dict[header[idx]] = cell_value
                # Only add non-empty rows (at least one non-empty value)
                if any(v for v in row_dict.values() if v):
                    data.append(row_dict)

        # CSV Upload
        elif ext == "csv":
            file_content = file.read().decode("utf-8")
            rows = csv.DictReader(file_content.splitlines())
            data = list(rows)

        else:
            return JsonResponse({"error": "Only CSV or Excel allowed"}, status=400)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"File parsing error: {str(e)}")
        print(error_details)
        return JsonResponse({"error": f"File parsing failed: {str(e)}"}, status=400)

    if not data:
        return JsonResponse({"error": "File is empty"}, status=400)

    # -------------------------------------------------------------------
    # Route to model-phase handlers (case-insensitive)
    # -------------------------------------------------------------------
    model_name_lower = model_name.lower()
    
    results_payload = {}
    row_results = [] # To be populated by handlers: [{'status': 'Success', 'error': ''}, ...]

    if model_name_lower == "itemmasterold":
        results_payload = handle_itemmaster_old_upload(data, request)
    elif model_name_lower in ("itemmaster", "material"):
        if phase == "1":
            results_payload = handle_itemmaster_phase_1(data, request)
        elif phase == "2":
            results_payload = handle_itemmaster_phase_2(data, request)
        else:
            return JsonResponse({"error": f"Invalid phase '{phase}' for ItemMaster. Use phase=1 or phase=2"}, status=400)
    elif model_name_lower == "matgattributeitem" or model_name == "MatgAttributeItem":
        if phase == "1":
            results_payload = handle_matgattribute_phase_1(data, request)
        else:
            return JsonResponse({"error": f"Invalid phase '{phase}' for MatgAttributeItem"}, status=400)
    elif model_name_lower == "matgroup":
        results_payload = handle_matgroup_upload(data, request)
    else:
        # Generic handler for all other models
        results_payload = handle_generic_model_upload(data, request, Model, model_name)

    # If the response is a JsonResponse (from handlers that haven't been refactored yet), 
    # extract the data if it contains detailed results.
    # New handlers should return a dict instead of JsonResponse internally.
    if isinstance(results_payload, JsonResponse):
        # Fallback for handlers not yet modified to support the new logging structure
        response_data = json.loads(results_payload.content.decode('utf-8'))
        row_results = response_data.get("row_results", [])
        if not row_results:
            # Generate dummy results from errors list if detailed row_results missing
            errors_map = {e.get('row'): e.get('error') for e in response_data.get('errors', [])}
            for i in range(len(data)):
                row_num = i + 2
                if row_num in errors_map:
                    row_results.append({'status': 'Error', 'error': errors_map[row_num]})
                else:
                    row_results.append({'status': 'Success', 'error': ''})
    else:
        # Expected structure from refactored handlers
        row_results = results_payload.get("row_results", [])
        response_data = results_payload

    # -------------------------------------------------------------------
    # Generate and Attach Log File if row_results exist
    # -------------------------------------------------------------------
    if row_results:
        try:
            wb = generate_upload_log(data, row_results, model_name)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            log_base64 = base64.b64encode(output.read()).decode('utf-8')
            response_data["log_file_base64"] = log_base64
            response_data["log_file_name"] = f"Upload_Log_{model_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        except Exception as e:
            print(f"Log generation failed: {str(e)}")

    if isinstance(results_payload, JsonResponse):
        # We need to rebuild the JsonResponse with the added log file
        return JsonResponse(response_data, status=results_payload.status_code)
    
    return JsonResponse(response_data)


def get_model_fields(request):
    model_name = request.GET.get("model")
    Model = get_model_by_name(model_name)

    if not Model:
        return JsonResponse({"error": "Invalid model"}, status=400)

    fields = [f.name for f in Model._meta.fields if f.name != "id"]
    return JsonResponse({"fields": fields})


# -------------------------------------------------------------------
# Generate ItemMaster Base Values Template
# -------------------------------------------------------------------
def generate_itemmaster_base_template(Model):
    """Generate template for ItemMaster base values with fixed column order and friendly headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Material Upload"

    # Fixed column definitions: (header label, field_key, width)
    columns = [
        ("SAP Item ID",     "sap_item_id",    15),
        ("Mat Type Code",   "mat_type_code",  15),
        ("Mgrp Code",       "mgrp_code",      15),
        ("Short Name",      "short_name",     25),
        ("Long Name",       "long_name",      35),
        ("Mgrp Long Name",  "mgrp_long_name", 30),
        ("SAP Name",        "sap_name",       25),
        ("Search Text",     "search_text",    30),
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Sample rows
    sample_data = [
        {
            "sap_item_id": "12345", "mat_type_code": "ROH", "mgrp_code": "MPFITLATR",
            "short_name": "SS LATROLET 3/4\"", "long_name": "Stainless Steel Latrolet 3/4 inch",
            "mgrp_long_name": "Pipe Fittings Latrolet", "sap_name": "SS LATROLET 3/4\"",
            "search_text": "latrolet pipe fitting ss",
        },
        {
            "sap_item_id": "12346", "mat_type_code": "ROH", "mgrp_code": "MPFITNIPH",
            "short_name": "CS NIPPLE HEX 1\"", "long_name": "Carbon Steel Hex Nipple 1 inch",
            "mgrp_long_name": "Pipe Fittings Nipple Hex", "sap_name": "CS NIPPLE HEX 1\"",
            "search_text": "nipple hex pipe fitting cs",
        },
    ]

    for row_idx, sample_row in enumerate(sample_data, start=2):
        for col_idx, (_, field_key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample_row.get(field_key, ""))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Material_Upload_Template.xlsx"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate ItemMaster Old Version Template (single combined sheet)
# -------------------------------------------------------------------
def generate_itemmaster_old_template():
    """Old-version single-sheet template: base fields + spec fields in one sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Material Old Version"

    columns = [
        ("SAP Item ID",   "sap_item_id",   15),
        ("Mat Type Code", "mat_type_code", 15),
        ("Mgrp Code",     "mgrp_code",     15),
        ("Item Desc",     "short_name",    30),
        ("Notes",         "long_name",     35),
        ("Search Text",   "search_text",   30),
        ("Type",          "item_type",     15),
        ("Number",        "item_number",   15),
        ("MOC",           "moc",           15),
        ("Size",          "item_size",     15),
        ("Part Number",   "part_number",   18),
        ("Model",         "model",         18),
        ("Make",          "make",          18),
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    sample_data = [
        {
            "sap_item_id": "12345", "mat_type_code": "ROH", "mgrp_code": "PIPES",
            "short_name": "SS PIPE SCH40 2\"", "long_name": "Stainless Steel Pipe SCH40 2 inch",
            "search_text": "pipe ss sch40", "item_type": "Seamless", "item_number": "P-001",
            "moc": "SS316", "item_size": "2 inch", "part_number": "", "model": "", "make": "Jindal",
        },
        {
            "sap_item_id": "12346", "mat_type_code": "ROH", "mgrp_code": "VALVES",
            "short_name": "CS GATE VALVE 1\"", "long_name": "Carbon Steel Gate Valve 1 inch",
            "search_text": "valve gate cs", "item_type": "Gate", "item_number": "V-002",
            "moc": "CS", "item_size": "1 inch", "part_number": "GV-001", "model": "GV100", "make": "L&T",
        },
    ]

    for row_idx, sample_row in enumerate(sample_data, start=2):
        for col_idx, (_, field_key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample_row.get(field_key, ""))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ItemMaster_OldVersion_template.xlsx"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Handler: ItemMaster Old Version Upload (single combined sheet)
# -------------------------------------------------------------------
def handle_itemmaster_old_upload(data, request):
    from itemmaster.models import ItemMaster
    from MaterialType.models import MaterialType
    from matgroups.models import MatGroup

    def get_value(row, keys):
        for k in keys:
            if k in row and str(row[k]).strip():
                return str(row[k]).strip()
        return None

    now = timezone.now()
    objs = []
    errors = []

    for idx, row in enumerate(data, start=2):
        try:
            sap_item_id_raw = get_value(row, ["SAP Item ID", "sap_item_id", "Sap Item Id", "SAP ITEM ID"])
            sap_item_id = int(sap_item_id_raw) if sap_item_id_raw and sap_item_id_raw.isdigit() else None

            mat_type_code_val = get_value(row, ["Mat Type Code", "mat_type_code", "MAT TYPE CODE"])
            mgrp_code_val = get_value(row, ["Mgrp Code", "mgrp_code", "MGRP CODE"])

            if not mat_type_code_val or not mgrp_code_val:
                errors.append({"row": idx, "error": "Mat Type Code and Mgrp Code are required"})
                continue

            mat_type_obj = MaterialType.objects.filter(mat_type_code=mat_type_code_val.upper()).first()
            mgrp_obj = MatGroup.objects.filter(mgrp_code=mgrp_code_val.upper()).first()

            if not mat_type_obj:
                errors.append({"row": idx, "error": f"MaterialType '{mat_type_code_val}' not found"})
                continue
            if not mgrp_obj:
                errors.append({"row": idx, "error": f"MatGroup '{mgrp_code_val}' not found"})
                continue

            short_name = get_value(row, ["Item Desc", "item_desc", "Short Name", "short_name"]) or ""
            long_name = get_value(row, ["Notes", "notes", "Long Name", "long_name"]) or ""
            search_text = get_value(row, ["Search Text", "search_text"]) or ""
            item_type = get_value(row, ["Type", "item_type"])
            item_number = get_value(row, ["Number", "item_number"])
            moc = get_value(row, ["MOC", "moc"])
            item_size = get_value(row, ["Size", "item_size"])
            part_number = get_value(row, ["Part Number", "part_number"])
            model = get_value(row, ["Model", "model"])
            make = get_value(row, ["Make", "make"])

            fields = dict(
                mat_type_code=mat_type_obj,
                mgrp_code=mgrp_obj,
                short_name=short_name,
                long_name=long_name,
                search_text=search_text,
                item_type=item_type,
                item_number=item_number,
                moc=moc,
                item_size=item_size,
                part_number=part_number,
                model=model,
                make=make,
                updated=now,
            )

            if sap_item_id:
                existing = ItemMaster.objects.filter(sap_item_id=sap_item_id).first()
                if existing:
                    for k, v in fields.items():
                        setattr(existing, k, v)
                    existing.save()
                    objs.append(existing)
                    continue

            objs.append(ItemMaster(sap_item_id=sap_item_id, created=now, **fields))

        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    new_objs = [o for o in objs if not o.pk]
    if new_objs:
        ItemMaster.objects.bulk_create(new_objs, ignore_conflicts=True)

    # Populate final row_results
    row_results = []
    errors_dict = {e['row']: e['error'] for e in errors}
    for i in range(len(data)):
        row_num = i + 2
        if row_num in errors_dict:
            row_results.append({'status': 'Error', 'error': errors_dict[row_num]})
        else:
            row_results.append({'status': 'Success', 'error': ''})

    return {
        "message": "ItemMaster Old Version upload complete",
        "inserted": len(objs),
        "errors": errors,
        "row_results": row_results
    }


# -------------------------------------------------------------------
# Generate MatgAttributeItem Template
# -------------------------------------------------------------------
def generate_matgattribute_template(Model):
    """Generate template for MatgAttributeItem with proper sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MatgAttributeItem"
    
    # Fields to exclude
    exclude_fields = {'id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted', 
                     'createdby_id', 'updatedby_id'}
    
    # Get data entry fields
    data_entry_fields = []
    for field in Model._meta.concrete_fields:
        if field.name not in exclude_fields and not field.many_to_many:
            data_entry_fields.append(field)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers
    headers = []
    for field in data_entry_fields:
        if hasattr(field, 'db_column') and field.db_column:
            field_display = field.db_column.replace('_', ' ').title()
        else:
            field_display = field.name.replace('_', ' ').title()
        headers.append(field_display)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
    
    # Add sample data specific to MatgAttributeItem
    sample_data = [
        {
            "mgrp_code": "GRP001",
            "attribute_name": "Color",
            "possible_values": "Red, Blue, Green, Yellow",
            "uom": "",
            "print_priority": "1",
            "validation": ""
        },
        {
            "mgrp_code": "GRP001",
            "attribute_name": "Size",
            "possible_values": "Small, Medium, Large, XL",
            "uom": "",
            "print_priority": "2",
            "validation": ""
        },
        {
            "mgrp_code": "GRP002",
            "attribute_name": "Weight",
            "possible_values": "1kg, 2kg, 5kg, 10kg",
            "uom": "kg",
            "print_priority": "1",
            "validation": ""
        },
    ]
    
    for row_idx, sample_row in enumerate(sample_data, start=2):
        for col_idx, field in enumerate(data_entry_fields, start=1):
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            value = sample_row.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "MatgAttributeItem_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate ItemMaster Attributes Template
# -------------------------------------------------------------------
def generate_itemmaster_attributes_template(mgrp_code=None):
    """Generate wide-format attributes template: Sap Item Id | Uom | <attr1> | <attr2> ..."""
    from matg_attributes.models import MatgAttributeItem
    from matgroups.models import MatGroup

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attribute Settings"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Fetch attribute names for the given mgrp_code (ordered by sequence)
    attr_names = []
    if mgrp_code:
        try:
            matgroup = MatGroup.objects.filter(mgrp_code=mgrp_code.strip().upper()).first()
            if matgroup:
                attr_names = list(
                    MatgAttributeItem.objects.filter(mgrp_code=matgroup, is_deleted=False)
                    .order_by("print_priority")
                    .values_list("attribute_name", flat=True)
                )
        except Exception:
            pass

    # Fixed columns + one column per attribute
    headers = ["Sap Item Id", "Uom"] + attr_names

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    # Two blank sample rows
    for row_idx in range(2, 4):
        ws.cell(row=row_idx, column=1, value="")
        ws.cell(row=row_idx, column=2, value="")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    mgrp_label = f"_{mgrp_code.upper()}" if mgrp_code else ""
    filename = f"ItemMaster_Attributes{mgrp_label}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate MatGroup Template (excludes attribgrpid, uom_values, notes)
# -------------------------------------------------------------------
def generate_matgroup_template():
    """Generate upload template for MatGroup with only required fields."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MatGroup"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Mgrp Code", "Sgrp Code", "Search Type", "Mgrp Shortname", "Mgrp Longname"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    sample_rows = [
        ["SAFTYAPRN", "SAFTY", "Materials", "Sample Mgrp Shortname 1", "Sample Mgrp Longname 1"],
        ["SAFTYHGVL", "SAFTY", "Materials", "Sample Mgrp Shortname 2", "Sample Mgrp Longname 2"],
    ]
    for row_idx, row in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="MatGroup_template.xlsx"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate Excel Template Dynamically
# -------------------------------------------------------------------
@csrf_exempt
def generate_excel_template(request):
    model_name = request.GET.get("model")
    template_type = request.GET.get("type", "base")  # "base" or "attributes" for ItemMaster
    
    if not model_name:
        return JsonResponse({"error": "Model name is required"}, status=400)
    
    Model = get_model_by_name(model_name)
    if not Model:
        # Try to get available models for debugging
        try:
            available_models = [m.__name__ for m in apps.get_models()]
            return JsonResponse({
                "error": f"Invalid model: {model_name}",
                "available_models": sorted(available_models)
            }, status=400)
        except:
            return JsonResponse({"error": f"Invalid model: {model_name}"}, status=400)
    
    # Old Version (single combined sheet)
    if model_name.lower() == "itemmasterold":
        return generate_itemmaster_old_template()

    # For ItemMaster or Material, handle separate downloads
    if model_name.lower() in ("itemmaster", "material"):
        from itemmaster.models import ItemMaster
        if template_type == "attributes":
            mgrp_code = request.GET.get("mgrp_code", "")
            return generate_itemmaster_attributes_template(mgrp_code=mgrp_code)
        else:
            return generate_itemmaster_base_template(ItemMaster)
    
    # Special handling for MatgAttributeItem - better sample data
    if Model.__name__.lower() == "matgattributeitem":
        return generate_matgattribute_template(Model)

    # Special handling for MatGroup - exclude attribgrpid, uom_values, notes
    if Model.__name__.lower() == "matgroup":
        return generate_matgroup_template()
    
    # Fields to exclude (audit fields)
    exclude_fields = {'id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted', 
                     'createdby_id', 'updatedby_id', 'local_item_id'}
    
    # Get all fields from the model
    all_fields = Model._meta.concrete_fields
    
    # Filter fields - only include data entry fields
    data_entry_fields = []
    for field in all_fields:
        if field.name not in exclude_fields and not field.many_to_many:
            data_entry_fields.append(field)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers
    headers = []
    for field in data_entry_fields:
        if hasattr(field, 'db_column') and field.db_column:
            field_display = field.db_column.replace('_', ' ').title()
        else:
            field_display = field.name.replace('_', ' ').title()
        headers.append(field_display)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    # Add sample data based on field types
    sample_rows = []
    for i in range(2):  # Add 2 sample rows
        sample_row = {}
        for field in data_entry_fields:
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            field_type = field.get_internal_type()
            
            if field_type == 'ForeignKey':
                sample_row[field_name] = f"SAMPLE_FK_{i+1}"
            elif field_type == 'IntegerField':
                sample_row[field_name] = 1000 + i
            elif field_type == 'BooleanField':
                sample_row[field_name] = "True" if i == 0 else "False"
            elif field_type == 'CharField':
                max_length = field.max_length if hasattr(field, 'max_length') else 50
                sample_row[field_name] = f"Sample {field_name.replace('_', ' ').title()} {i+1}"[:max_length]
            elif field_type == 'JSONField':
                # For JSONField, provide comma-separated values as string
                sample_row[field_name] = "Value1, Value2, Value3"
            else:
                sample_row[field_name] = f"Sample Value {i+1}"
        
        sample_rows.append(sample_row)
    
    # Write sample data
    for row_idx, sample_row in enumerate(sample_rows, start=2):
        for col_idx, field in enumerate(data_entry_fields, start=1):
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            value = sample_row.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model_name}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model_name}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
